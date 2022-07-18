from csv import excel
from openpyxl import load_workbook, workbook

arquivo_etiqueta = load_workbook('Relação.xlsx')
print(arquivo_etiqueta)

bade_dados = arquivo_etiqueta.active
etiqueta = arquivo_etiqueta.create_sheet("Etiqueta")

max_linha = bade_dados.max_row
max_coluna = bade_dados.max_column

for linha in range(1, max_linha + 1):

    for coluna in range(1,max_coluna + 1):     
        celula = bade_dados.cell(row=linha, column= coluna)
        etiqueta.cell(row=linha, column= coluna, value = celula.value)


controle = 2
for linha in range(1,max_linha + 1):
    
    if linha == 1:
        etiqueta.insert_rows(idx=controle, amount= 2 )
    else:
        etiqueta.insert_rows(idx= controle, amount=2 )
    controle += 3

controle = 1

for linha in range(1, max_linha + 1):

    for coluna in range(1,max_coluna + 1):  
        if   linha == 1:

            celula = etiqueta.cell(row=linha, column= 5)
            etiqueta.cell(row=linha + 1, column= 2, value = celula.value)
        else:

            celula = etiqueta.cell(row=controle, column= 5)
            etiqueta.cell(row=controle + 1, column= 2, value = celula.value)

    controle += 3


controle = 1
for linha in range(1, max_linha + 1):

    for coluna in range(1,max_coluna + 1):  
        if   linha == 1:

            celula = etiqueta.cell(row=linha, column= 6)
            etiqueta.cell(row=linha + 1, column= 3, value = celula.value)
        else:
            celula = etiqueta.cell(row=controle , column= 6)
            etiqueta.cell(row=controle + 1, column= 3, value = celula.value)

    controle += 3

etiqueta.delete_cols(idx= 5)
etiqueta.delete_cols(idx= 5)


arquivo_etiqueta.save('Etiqueta.xlsx')
